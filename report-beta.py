import openpyxl, sys, os
from dateutil.parser import parse
from datetime import timedelta
from datetime import datetime
def sort_report(generated_report):
    wb = openpyxl.load_workbook(generated_report)
    sheet1 = wb['Sheet1']
    #auto pick first date and decide next four days of the week.
    next_day = timedelta(days=1)
    day_one = sheet1.cell(row = 2, column=1).value
    day_two = day_one + next_day
    day_three = day_two + next_day
    day_four = day_three + next_day
    day_five = day_four + next_day

    provider_value = {}

    #create new sheets for each provider
    for i in range (2,20):
        wb.create_sheet(index=i, title='Sheet' + str(i))
    
    wb['Sheet2'].title = '21st Century'
    sheet2 = wb['21st Century']

    for row in range(2, sheet1.max_row + 1):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == '21st Century':
            provider_value[sheet1.cell(row = row, column = 2).value] = sheet1.cell(row = row, column = 4).value
        for col in  range(2,5):
            sheet2.cell(row = row - 1, column = col - 1).value = sheet1.cell(row = row, column = col).value
    for row in range(2, sheet1.max_row + 1):
        if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == '21st Century':
            provider_value[sheet1.cell(row = row, column = 2).value] = sheet1.cell(row = row, column = 4).value
            sheet2.cell(row = row - 1, column = col - 1).value = sheet1.cell(row = row, column = col).value
    for row in range(2, sheet1.max_row + 1):
        if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == '21st Century':
            provider_value[sheet1.cell(row = row, column = 2).value] = sheet1.cell(row = row, column = 4).value

    for row in range(2, sheet1.max_row + 1):
        if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == '21st Century':
            provider_value[sheet1.cell(row = row, column = 2).value] = sheet1.cell(row = row, column = 4).value

    for row in range(2, sheet1.max_row + 1):
        if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == '21st Century':
            provider_value[sheet1.cell(row = row, column = 2).value] = sheet1.cell(row = row, column = 4).value
        
sort_report('sept_demo.xlsx')