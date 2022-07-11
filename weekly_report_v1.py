import openpyxl
from dateutil.parser import parse
from datetime import timedelta
from datetime import datetime
import sys,os
generated_report = input('Enter the filename/path): ')
try:
    wb = openpyxl.load_workbook(generated_report)
except openpyxl.utils.exceptions.InvalidFileException:
    print('File format not supported. Supported formats are: .xlsx,.xlsm,.xltx,.xltm. Please try again with the correct file format.')
    sys.exit()
sheet1 = wb['Sheet1']
#cleaning the data
for row in range(1, sheet1.max_row + 1):
    for column in range(1, sheet1.max_column + 1):
        if sheet1.cell(row = row, column = 2).value == 'IDOWU-TAYLOR' or \
            sheet1.cell(row = row, column = 2).value == 'Interswitch_IPNX_Interface' or \
            sheet1.cell(row = row, column = 2).value == 'Interswitch_MainOne_Interface' or \
            sheet1.cell(row = row, column = column).value == None:
            sheet1.delete_rows(row)
#convert the dates from string to date format
dates = []
for row in range(2, sheet1.max_row + 1):
    dates.append(sheet1.cell(row = row, column=1).value)
for i in range(len(dates)):
    new_date = parse(dates[i])
    new_date = new_date.date()
    dates[i] = new_date
#return the converted date into the date column
for x in range(len(dates)):
    sheet1.cell(row = x  + 2, column = 1).value = dates[x]
#auto pick first date and decide next four days of the week.
next_day = timedelta(days=1)
day_one = sheet1.cell(row = 2, column=1).value
day_two = day_one + next_day
day_three = day_two + next_day
day_four = day_three + next_day
day_five = day_four + next_day
#remove % and '.00' if value is 100.00 or 0.00 (changing values from percentage to number)
raw_values = []
for row in range(2, sheet1.max_row + 1):
    raw_values.append(sheet1.cell(row = row, column = 4).value)
for i in range(len(raw_values)):
    if raw_values[i] != None:
        new_value = raw_values[i].removesuffix(' %')
        raw_values[i] = float(new_value)
    if raw_values[i] == '100.00' or raw_values[i] == '0.00':
        final_value = raw_values[i].removesuffix('.00')
        raw_values[i] = int(final_value)
#return the converted values into the values column
for x in range(len(raw_values)):
    sheet1.cell(row = x  + 2, column = 4).value = raw_values[x]
#wb.save('In_progress.xlsx')
#create new sheets for each provider
for i in range (2,20):
    wb.create_sheet(index=i, title='Sheet' + str(i))
#copy data for each provider
#21st Century
wb['Sheet2'].title = '21st Century'
sheet2 = wb['21st Century']
#empty list to store value for other weekdays
day2_21st = []
day3_21st = []
day4_21st = []
day5_21st = []
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == '21st Century':
            sheet2.cell(row = row - 1, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == '21st Century':
        day2_21st.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_21st)):
            sheet2.cell(row = x + 1, column = 4).value = day2_21st[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == '21st Century':
        day3_21st.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_21st)):
            sheet2.cell(row = x + 1, column = 5).value = day3_21st[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == '21st Century':
        day4_21st.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_21st)):
            sheet2.cell(row = x + 1, column = 6).value = day4_21st[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == '21st Century':
        day5_21st.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_21st)):
            sheet2.cell(row = x + 1, column = 7).value = day5_21st[x]
#Airtel
wb['Sheet3'].title = 'Airtel'
sheet3 = wb['Airtel']
#empty list to store value for other weekdays
day2_airtel = []
day3_airtel = []
day4_airtel = []
day5_airtel = []
airtel_row = sheet2.max_row + 1
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Airtel':
            sheet3.cell(row = row - airtel_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Airtel':
        day2_airtel.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_airtel)):
            sheet3.cell(row = x + 1, column = 4).value = day2_airtel[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Airtel':
        day3_airtel.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_airtel)):
            sheet3.cell(row = x + 1, column = 5).value = day3_airtel[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Airtel':
        day4_airtel.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_airtel)):
            sheet3.cell(row = x + 1, column = 6).value = day4_airtel[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Airtel':
        day5_airtel.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_airtel)):
            sheet3.cell(row = x + 1, column = 7).value = day5_airtel[x]
#Broadbased
wb['Sheet4'].title = 'Broadbased'
sheet4 = wb['Broadbased']
#empty list to store value for other weekdays
day2_broadbased = []
day3_broadbased = []
day4_broadbased = []
day5_broadbased = []
broadbased_row = sheet3.max_row + airtel_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Broadbased':
            sheet4.cell(row = row - broadbased_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Broadbased':
        day2_broadbased.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_broadbased)):
            sheet4.cell(row = x + 1, column = 4).value = day2_broadbased[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Broadbased':
        day3_broadbased.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_broadbased)):
            sheet4.cell(row = x + 1, column = 5).value = day3_broadbased[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Broadbased':
        day4_broadbased.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_broadbased)):
            sheet4.cell(row = x + 1, column = 6).value = day4_broadbased[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Broadbased':
        day5_broadbased.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_broadbased)):
            sheet4.cell(row = x + 1, column = 7).value = day5_broadbased[x]
#Cyberspace
wb['Sheet5'].title = 'Cyberspace'
sheet5 = wb['Cyberspace']
#empty list to store value for other weekdays
day2_cyberspace = []
day3_cyberspace = []
day4_cyberspace = []
day5_cyberspace = []
cyber_row = sheet4.max_row + broadbased_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Cyberspace':
            sheet5.cell(row = row - cyber_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Cyberspace':
        day2_cyberspace.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_cyberspace)):
            sheet5.cell(row = x + 1, column = 4).value = day2_cyberspace[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Cyberspace':
        day3_cyberspace.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_cyberspace)):
            sheet5.cell(row = x + 1, column = 5).value = day3_cyberspace[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Cyberspace':
        day4_cyberspace.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_cyberspace)):
            sheet5.cell(row = x + 1, column = 6).value = day4_cyberspace[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Cyberspace':
        day5_cyberspace.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_cyberspace)):
            sheet5.cell(row = x + 1, column = 7).value = day5_cyberspace[x]
wb.save('test.xlsx')
#DCC
wb['Sheet18'].title = 'DCC'
sheet18 = wb['DCC']
#empty list to store value for other weekdays
day2_dcc = []
day3_dcc = []
day4_dcc = []
day5_dcc = []
dcc_row = sheet5.max_row + cyber_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'DCC':
            if row == dcc_row:
                dcc_row = dcc_row - 1
            else:
                dcc_row = dcc_row
            sheet18.cell(row = row - dcc_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'DCC':
        day2_dcc.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_dcc)):
            sheet18.cell(row = x + 1, column = 4).value = day2_dcc[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'DCC':
        day3_dcc.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_dcc)):
            sheet18.cell(row = x + 1, column = 5).value = day3_dcc[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'DCC':
        day4_dcc.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_dcc)):
            sheet18.cell(row = x + 1, column = 6).value = day4_dcc[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'DCC':
        day5_dcc.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_dcc)):
            sheet18.cell(row = x + 1, column = 7).value = day5_dcc[x]
#EMP_Egypt
wb['Sheet19'].title = 'EMP'
sheet19 = wb['EMP']
#empty list to store value for other weekdays
day2_emp = []
day3_emp = []
day4_emp = []
day5_emp = []
emp_row = sheet18.max_row + dcc_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'EMP':
            sheet19.cell(row = row - emp_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'EMP':
        day2_emp.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_emp)):
            sheet19.cell(row = x + 1, column = 4).value = day2_emp[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'EMP':
        day3_emp.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_emp)):
            sheet19.cell(row = x + 1, column = 5).value = day3_emp[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'EMP':
        day4_emp.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_emp)):
            sheet19.cell(row = x + 1, column = 6).value = day4_emp[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'EMP':
        day5_emp.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_emp)):
            sheet19.cell(row = x + 1, column = 7).value = day5_emp[x]
#E-STREAM
wb['Sheet6'].title = 'E-STREAM'
sheet6 = wb['E-STREAM']
#empty list to store value for other weekdays
day2_estream = []
day3_estream = []
day4_estream = []
day5_estream = []
estream_row = sheet19.max_row + emp_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'E-STREAM':
            sheet6.cell(row = row - estream_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'E-STREAM':
        day2_estream.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_estream)):
            sheet6.cell(row = x + 1, column = 4).value = day2_estream[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'E-STREAM':
        day3_estream.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_estream)):
            sheet6.cell(row = x + 1, column = 5).value = day3_estream[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'E-STREAM':
        day4_estream.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_estream)):
            sheet6.cell(row = x + 1, column = 6).value = day4_estream[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'E-STREAM':
        day5_estream.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_estream)):
            sheet6.cell(row = x + 1, column = 7).value = day5_estream[x]
#Globacom
wb['Sheet7'].title = 'Globacom'
sheet7 = wb['Globacom']
#empty list to store value for other weekdays
day2_glo = []
day3_glo = []
day4_glo = []
day5_glo = []
glo_row = sheet6.max_row + estream_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Globacom':
            sheet7.cell(row = row - glo_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Globacom':
        day2_glo.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_glo)):
            sheet7.cell(row = x + 1, column = 4).value = day2_glo[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Globacom':
        day3_glo.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_glo)):
            sheet7.cell(row = x + 1, column = 5).value = day3_glo[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Globacom':
        day4_glo.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_glo)):
            sheet7.cell(row = x + 1, column = 6).value = day4_glo[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Globacom':
        day5_glo.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_glo)):
            sheet7.cell(row = x + 1, column = 7).value = day5_glo[x]
#ICSL
wb['Sheet8'].title = 'ICSL'
sheet8 = wb['ICSL']
#empty list to store value for other weekdays
day2_icsl = []
day3_icsl = []
day4_icsl = []
day5_icsl = []
icsl_row = sheet7.max_row + glo_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'ICSL':
            sheet8.cell(row = row - icsl_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'ICSL':
        day2_icsl.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_icsl)):
            sheet8.cell(row = x + 1, column = 4).value = day2_icsl[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'ICSL':
        day3_icsl.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_icsl)):
            sheet8.cell(row = x + 1, column = 5).value = day3_icsl[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'ICSL':
        day4_icsl.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_icsl)):
            sheet8.cell(row = x + 1, column = 6).value = day4_icsl[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'ICSL':
        day5_icsl.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_icsl)):
            sheet8.cell(row = x + 1, column = 7).value = day5_icsl[x]
#IPNX
wb['Sheet9'].title = 'IPNX'
sheet9 = wb['IPNX']
#empty list to store value for other weekdays
day2_ipnx = []
day3_ipnx = []
day4_ipnx = []
day5_ipnx = []
ipnx_row = sheet8.max_row + icsl_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'IPNX':
            sheet9.cell(row = row - ipnx_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'IPNX':
        day2_ipnx.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_ipnx)):
            sheet9.cell(row = x + 1, column = 4).value = day2_ipnx[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'IPNX':
        day3_ipnx.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_ipnx)):
            sheet9.cell(row = x + 1, column = 5).value = day3_ipnx[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'IPNX':
        day4_ipnx.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_ipnx)):
            sheet9.cell(row = x + 1, column = 6).value = day4_ipnx[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'IPNX':
        day5_ipnx.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_ipnx)):
            sheet9.cell(row = x + 1, column = 7).value = day5_ipnx[x]
#MainOne
wb['Sheet10'].title = 'MainOne'
sheet10 = wb['MainOne']
#empty list to store value for other weekdays
day2_mainone = []
day3_mainone = []
day4_mainone = []
day5_mainone = []
mainone_row = sheet9.max_row + ipnx_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'MainOne':
            sheet10.cell(row = row - mainone_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'MainOne':
        day2_mainone.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_mainone)):
            sheet10.cell(row = x + 1, column = 4).value = day2_mainone[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'MainOne':
        day3_mainone.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_mainone)):
            sheet10.cell(row = x + 1, column = 5).value = day3_mainone[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'MainOne':
        day4_mainone.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_mainone)):
            sheet10.cell(row = x + 1, column = 6).value = day4_mainone[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'MainOne':
        day5_mainone.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_mainone)):
            sheet10.cell(row = x + 1, column = 7).value = day5_mainone[x]
#MTN
wb['Sheet11'].title = 'MTN'
sheet11 = wb['MTN']
#empty list to store value for other weekdays
day2_mtn = []
day3_mtn = []
day4_mtn = []
day5_mtn = []
mtn_row = sheet10.max_row + mainone_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'MTN':
            sheet11.cell(row = row - mtn_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'MTN':
        day2_mtn.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_mtn)):
            sheet11.cell(row = x + 1, column = 4).value = day2_mtn[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'MTN':
        day3_mtn.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_mtn)):
            sheet11.cell(row = x + 1, column = 5).value = day3_mtn[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'MTN':
        day4_mtn.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_mtn)):
            sheet11.cell(row = x + 1, column = 6).value = day4_mtn[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'MTN':
        day5_mtn.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_mtn)):
            sheet11.cell(row = x + 1, column = 7).value = day5_mtn[x]
#Priority
wb['Sheet12'].title = 'Priority'
sheet12 = wb['Priority']
#empty list to store value for other weekdays
day2_priority = []
day3_priority = []
day4_priority = []
day5_priority = []
priority_row = sheet11.max_row + mtn_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Priority':
            sheet12.cell(row = row - priority_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Priority':
        day2_priority.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_priority)):
            sheet12.cell(row = x + 1, column = 4).value = day2_priority[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Priority':
        day3_priority.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_priority)):
            sheet12.cell(row = x + 1, column = 5).value = day3_priority[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Priority':
        day4_priority.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_priority)):
            sheet12.cell(row = x + 1, column = 6).value = day4_priority[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Priority':
        day5_priority.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_priority)):
            sheet12.cell(row = x + 1, column = 7).value = day5_priority[x]

#Shivtech
wb['Sheet13'].title = 'Shivtech'
sheet13 = wb['Shivtech']
#empty list to store value for other weekdays
day2_shivtech = []
day3_shivtech = []
day4_shivtech = []
day5_shivtech = []
shivtech_row = sheet12.max_row + priority_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Shivtech':
            sheet13.cell(row = row - shivtech_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Shivtech':
        day2_shivtech.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_shivtech)):
            sheet13.cell(row = x + 1, column = 4).value = day2_shivtech[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Shivtech':
        day3_shivtech.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_shivtech)):
            sheet13.cell(row = x + 1, column = 5).value = day3_shivtech[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Shivtech':
        day4_shivtech.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_shivtech)):
            sheet13.cell(row = x + 1, column = 6).value = day4_shivtech[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Shivtech':
        day5_shivtech.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_shivtech)):
            sheet13.cell(row = x + 1, column = 7).value = day5_shivtech[x]

#Swift Networks
wb['Sheet14'].title = 'Swift Networks'
sheet14 = wb['Swift Networks']
#empty list to store value for other weekdays
day2_swiftng = []
day3_swiftng = []
day4_swiftng = []
day5_swiftng = []
swiftng_row = sheet13.max_row + shivtech_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Swift Networks':
            sheet14.cell(row = row - swiftng_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Swift Networks':
        day2_swiftng.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_swiftng)):
            sheet14.cell(row = x + 1, column = 4).value = day2_swiftng[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Swift Networks':
        day3_swiftng.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_swiftng)):
            sheet14.cell(row = x + 1, column = 5).value = day3_swiftng[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Swift Networks':
        day4_swiftng.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_swiftng)):
            sheet14.cell(row = x + 1, column = 6).value = day4_swiftng[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Swift Networks':
        day5_swiftng.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_swiftng)):
            sheet14.cell(row = x + 1, column = 7).value = day5_swiftng[x]
#SwiftTalk
wb['Sheet15'].title = 'SwiftTalk'
sheet15 = wb['SwiftTalk']
#empty list to store value for other weekdays
day2_swifttalk = []
day3_swifttalk = []
day4_swifttalk = []
day5_swifttalk = []
swifttalk_row = sheet14.max_row + swiftng_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'SwiftTalk':
            sheet15.cell(row = row - swifttalk_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'SwiftTalk':
        day2_swifttalk.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_swifttalk)):
            sheet15.cell(row = x + 1, column = 4).value = day2_swifttalk[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'SwiftTalk':
        day3_swifttalk.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_swifttalk)):
            sheet15.cell(row = x + 1, column = 5).value = day3_swifttalk[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'SwiftTalk':
        day4_swifttalk.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_swifttalk)):
            sheet15.cell(row = x + 1, column = 6).value = day4_swifttalk[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'SwiftTalk':
        day5_swifttalk.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_swifttalk)):
            sheet15.cell(row = x + 1, column = 7).value = day5_swifttalk[x]
#Vodacom
wb['Sheet16'].title = 'Vodacom'
sheet16 = wb['Vodacom']
#empty list to store value for other weekdays
day2_vodacom = []
day3_vodacom = []
day4_vodacom = []
day5_vodacom = []
vodacom_row = sheet15.max_row + swifttalk_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Vodacom':
            sheet16.cell(row = row - vodacom_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Vodacom':
        day2_vodacom.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_vodacom)):
            sheet16.cell(row = x + 1, column = 4).value = day2_vodacom[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Vodacom':
        day3_vodacom.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_vodacom)):
            sheet16.cell(row = x + 1, column = 5).value = day3_vodacom[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Vodacom':
        day4_vodacom.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_vodacom)):
            sheet16.cell(row = x + 1, column = 6).value = day4_vodacom[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Vodacom':
        day5_vodacom.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_vodacom)):
            sheet16.cell(row = x + 1, column = 7).value = day5_vodacom[x]
#Vodacom_Offiste
wb['Sheet17'].title = 'Vodacom_Offsite'
sheet17 = wb['Vodacom_Offsite']
#empty list to store value for other weekdays
day2_vodacom_offsite = []
day3_vodacom_offsite = []
day4_vodacom_offsite = []
day5_vodacom_offsite = []
vodacomoffiste_row = sheet16.max_row + vodacom_row
for row in range(2, sheet1.max_row + 1):
    for col in range(2,5):
        if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 3).value == 'Vodacom_Offsite':
            sheet17.cell(row = row - vodacomoffiste_row, column = col - 1).value = sheet1.cell(row = row, column = col).value
for row in range(2, sheet1.max_row + 1):
    if sheet1.cell(row = row, column = 1).value == day_two and sheet1.cell(row = row, column = 3).value == 'Vodacom_Offsite':
        day2_vodacom_offsite.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day2_vodacom_offsite)):
            sheet17.cell(row = x + 1, column = 4).value = day2_vodacom_offsite[x]
    if sheet1.cell(row = row, column = 1).value == day_three and sheet1.cell(row = row, column = 3).value == 'Vodacom_Offsite':
        day3_vodacom_offsite.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day3_vodacom_offsite)):
            sheet17.cell(row = x + 1, column = 5).value = day3_vodacom_offsite[x]
    if sheet1.cell(row = row, column = 1).value == day_four and sheet1.cell(row = row, column = 3).value == 'Vodacom_Offsite':
        day4_vodacom_offsite.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day4_vodacom_offsite)):
            sheet17.cell(row = x + 1, column = 6).value = day4_vodacom_offsite[x]
    if sheet1.cell(row = row, column = 1).value == day_five and sheet1.cell(row = row, column = 3).value == 'Vodacom_Offsite':
        day5_vodacom_offsite.append(sheet1.cell(row = row, column = 4).value)
        for x in range(len(day5_vodacom_offsite)):
            sheet17.cell(row = x + 1, column = 7).value = day5_vodacom_offsite[x]
wb.save(filename='C:\\Users\\\'Lamide Ilori\\Desktop\\Work\\Report\\' + generated_report.removesuffix('.xlsx')\
    + '_sorted_on_' + datetime.now().strftime('%Y_%m_%d_%H_%M_%S') + '.xlsx')
print('Report modified and successfully saved.')
wb.close()